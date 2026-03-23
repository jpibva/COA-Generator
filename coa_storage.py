"""Persistencia compatible para configuración, historial, registro y sesión."""

import json
import logging
import os
import re
import shutil
from datetime import datetime

from coa_formats import (
    DEFAULT_BRAND_MAP,
    DEFAULT_CLIENT_MAP,
    DEFAULT_MICRO_FORMATS,
    DEFAULT_QUALITY_FORMATS,
)


CONFIG_FILE = "config.json"
MICRO_HISTORY = "Historial_Microbiologia.xlsx"
COA_REGISTRY = "Registro_COAs.xlsx"
SESSION_FILE = "session.json"
BACKUP_DIRNAME = "_backups"


def _ensure_parent_dir(path):
    parent = os.path.dirname(os.path.abspath(path))
    if parent:
        os.makedirs(parent, exist_ok=True)


def _create_backup(path, keep=5):
    """Crea un backup rotativo del archivo si ya existe."""
    if not os.path.exists(path):
        return None

    abs_path = os.path.abspath(path)
    parent = os.path.dirname(abs_path)
    filename = os.path.basename(abs_path)
    stem, ext = os.path.splitext(filename)
    backup_dir = os.path.join(parent, BACKUP_DIRNAME)
    os.makedirs(backup_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = os.path.join(backup_dir, f"{stem}_{timestamp}{ext}.bak")
    shutil.copy2(abs_path, backup_path)

    pattern = re.compile(rf"^{re.escape(stem)}_\d{{8}}_\d{{6}}_\d{{6}}{re.escape(ext)}\.bak$")
    backups = sorted(
        (
            os.path.join(backup_dir, name)
            for name in os.listdir(backup_dir)
            if pattern.match(name)
        ),
        key=os.path.getmtime,
        reverse=True,
    )
    for old_backup in backups[keep:]:
        try:
            os.remove(old_backup)
        except OSError:
            pass

    return backup_path


def load_config(config_file=CONFIG_FILE):
    defaults = {
        "pdf_folder": "C:/COA Generator/PL",
        "template_folder": "C:/COA Generator/Templates",
        "output_folder": "C:/Users/jibarra/Mi unidad/1. CoAs/1. EXPORTACIÓN",
        "micro_formats": DEFAULT_MICRO_FORMATS,
        "quality_formats": DEFAULT_QUALITY_FORMATS,
        "client_map": DEFAULT_CLIENT_MAP,
        "brand_map": DEFAULT_BRAND_MAP,
    }
    if os.path.exists(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                for key in defaults:
                    if key not in data:
                        data[key] = defaults[key]
                for fname, fdata in DEFAULT_MICRO_FORMATS.items():
                    if fname not in data["micro_formats"]:
                        data["micro_formats"][fname] = fdata
                for fname, fdata in DEFAULT_QUALITY_FORMATS.items():
                    if fname not in data["quality_formats"]:
                        data["quality_formats"][fname] = fdata
                if "client_map" not in data:
                    data["client_map"] = DEFAULT_CLIENT_MAP
                if "brand_map" not in data:
                    data["brand_map"] = DEFAULT_BRAND_MAP
                return data
        except Exception:
            pass
    return defaults


def save_config(cfg, config_file=CONFIG_FILE):
    _ensure_parent_dir(config_file)
    _create_backup(config_file)
    with open(config_file, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=4, ensure_ascii=False)


def normalize_product_name(nombre):
    return re.sub(r'\b\d+\s*(?:oz|g|kg|lb|ml|l)\b', '', nombre, flags=re.IGNORECASE).strip().lower()


def load_micro_history(history_file=MICRO_HISTORY):
    """Carga historial de TODAS las hojas del Excel (una por formato)."""
    history = {}
    if not os.path.exists(history_file):
        return history
    try:
        import openpyxl

        wb = openpyxl.load_workbook(history_file)
        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue
            headers = [cell.value for cell in ws[1]]
            if not headers or headers[0] is None:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                record = dict(zip(headers, row))
                key = (
                    str(record.get("Cliente", "")).lower().strip(),
                    normalize_product_name(str(record.get("Producto", ""))),
                    str(record.get("Lote", "")).strip(),
                )
                if key in history:
                    for k, v in record.items():
                        if v and not history[key].get(k):
                            history[key][k] = v
                else:
                    history[key] = record
    except Exception as e:
        logging.error(f"Error cargando historial micro: {e}")
    return history


def save_micro_history_record(cliente, producto, lote, micro_values, config, formato_nombre="", history_file=MICRO_HISTORY):
    """Guarda en una hoja separada por formato microbiológico."""
    try:
        import openpyxl

        fmt = config["micro_formats"].get(formato_nombre, {})
        params = fmt.get("parametros", [])
        tipo = fmt.get("tipo", "simple")
        claves = []
        for p in params:
            clave = p["clave"]
            if tipo == "n_series" and p.get("n_series", False):
                for n in range(1, 6):
                    claves.append(f"{clave}_n{n}")
            else:
                claves.append(clave)
        if not claves:
            claves = sorted(micro_values.keys())

        headers = ["Cliente", "Producto", "Lote"] + claves + ["Fecha"]
        nombre_limpio = re.sub(r'[\\/*?:\[\]]', '-', formato_nombre or "General")
        sheet_name = nombre_limpio[:31]

        if os.path.exists(history_file):
            wb = openpyxl.load_workbook(history_file)
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)

        prod_norm = normalize_product_name(producto)
        cliente_l = cliente.lower().strip()
        row_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if (
                str(row[0] or "").lower().strip() == cliente_l
                and normalize_product_name(str(row[1] or "")) == prod_norm
                and str(row[2] or "").strip() == lote
            ):
                row_idx = i
                break

        nueva_fila = [cliente, producto, lote]
        for k in claves:
            nueva_fila.append(micro_values.get(k, ""))
        nueva_fila.append(datetime.now().strftime("%d/%m/%Y %H:%M"))

        if row_idx:
            for col, val in enumerate(nueva_fila, start=1):
                ws.cell(row=row_idx, column=col, value=val)
        else:
            ws.append(nueva_fila)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        _ensure_parent_dir(history_file)
        _create_backup(history_file)
        wb.save(history_file)
    except Exception as e:
        logging.error(f"Error guardando historial micro: {e}")


def get_registro_path(base_dir):
    """Registro siempre en la carpeta del programa, junto a config.json."""
    return os.path.join(base_dir, COA_REGISTRY)


def registrar_coa(filas, base_dir):
    """Registro de COAs — una hoja por año, siempre en carpeta del programa."""
    try:
        import openpyxl

        registro_path = get_registro_path(base_dir)
        headers = [
            "Fecha Generación",
            "Embarque",
            "Cliente",
            "PO",
            "Producto",
            "Lotes",
            "Cajas Totales",
            "Archivo Generado",
        ]
        año_actual = str(datetime.now().year)

        if os.path.exists(registro_path):
            try:
                wb = openpyxl.load_workbook(registro_path)
            except Exception:
                registro_path = registro_path.replace(".xlsx", f"_{datetime.now().strftime('%H%M%S')}.xlsx")
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if año_actual in wb.sheetnames:
            ws = wb[año_actual]
        else:
            ws = wb.create_sheet(title=año_actual)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="2F5496")

        for fila in filas:
            ws.append(fila)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        _ensure_parent_dir(registro_path)
        _create_backup(registro_path)
        wb.save(registro_path)
    except Exception as e:
        logging.error(f"Error guardando registro COAs: {e}")


def get_session_path(base_dir):
    return os.path.join(base_dir, SESSION_FILE)


def save_session_data(session, base_dir):
    session_path = get_session_path(base_dir)
    _ensure_parent_dir(session_path)
    _create_backup(session_path)
    with open(session_path, "w", encoding="utf-8") as f:
        json.dump(session, f, indent=2, ensure_ascii=False)


def load_session_data(base_dir):
    spath = get_session_path(base_dir)
    if os.path.exists(spath):
        with open(spath, "r", encoding="utf-8") as f:
            return json.load(f)
    return None
